descrizione = '''
Programma: speedyPizza.py
Autore: Simone Tempesta
                                DESCRIZIONE: 
Inserita una directory contenente un file.xlsx(excel), con una formattazione predeterminata,
costruisce e riordina il file in un dataframe. Eseguita l'elaborazione, in output (nella directory 
selezionata) crea una cartella denominata (ordini_Y-M-D h#m#s) ed all'interno posiziona il 
file.xlsx(excel) elaborato denominato(ordine_h#m). Se si spunta la checkbox archivia originale, 
viene creata un'altra cartella denominata(archivioY-M-D) al cui interno viene posizionato il 
file originale(precedentemente elaborato) denominato(D-hfile.xlsx). Se si spunta la checkbox 
invia email, sarà inoltrato il file.xlsx elaborato.
                                
                                USO:
1)Premi il pulsante Seleziona e sciegli la cartella dove si trova il file.xlsx da elaborare
    (solo un excel potrà essere all'interno della cartella).
2)Seleziona con una spunta: @-Invia Email ed Archivia Originale. Verrà inviata una mail 
    al ristoratore ed il file originale sarà archiviato in una nuova cartella che si
    chiamerà: archivioAnno-Mese-Giorno, rinominato come Ordine_delle_ore#minuti.xlsx, 
    inoltre l'originale sarà eliminato dalla cartella dove si trova.
3)Premi il pulsante Avvia->, verrà eseguito il processo descritto sopra(in base alle spunte sulle 
    cheeckbox) e sarà creata una nuova cartella chiamata ordini_anno_mese_giorno ora#minuti#secondi 
    nella quale si troverà il file elaborato per l'ordine.
4) Al termine comparirà se tutto è andato a buon fine il messaggio: operazione avvenuta con successo.
'''
#Gestire errori di connessione per l'invio della mail
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill
import os
import smtplib
from email.message import EmailMessage
from datetime import datetime

#Scansiona il percorso
def leggi_directory(percorso):
    #directory = "C:\\Users\\simon\\PycharmProjects\\lavorare con exel"
    l_directory = os.listdir(percorso)
    return l_directory
###
#Estrai file xlsx
def estrai_file(lista_risorse):
    lista_file = []
    for elemnt in lista_risorse:
        if ".xlsx" in elemnt and ".~" not in elemnt :
            lista_file.append(elemnt)
    return lista_file

#Elabora file excel
def elaborazione_exel(lista_exel, percorso):
    for element in lista_exel:
        df_elaborato = elabora_exel(element, percorso)
        if "Informazioni cronologiche" in df_elaborato.columns:
            df_elaborato.drop("Informazioni cronologiche", axis=1, inplace=True)
        if "Indirizzo email" in df_elaborato.columns:
            df_elaborato.drop("Indirizzo email", axis=1, inplace=True)
    for column in df_elaborato.columns:
        column_not_number = ["Plesso", "Classe", "Informazioni cronologiche", "Indirizzo email"]
        if column not in column_not_number:
            df_elaborato.astype({column:int})
    df_elaborato.sort_values(by=["Plesso", "Classe"]).set_index("Classe", inplace=True)
    return df_elaborato

#Elabora Excel to DataFrame
def elabora_exel(file_excel, percorso):
    path_file = percorso+"/"+file_excel
    df_elaborato = pd.read_excel(path_file)
    #df_elaborato.dropna(axis=1, how="all", inplace=True)
    df_elaborato.fillna(value=0, inplace=True)
    return df_elaborato

#Elabora DataFrame
def elabora_df_elab(df_elaborato):
    classi = df_elaborato.groupby(by=["Plesso","Classe"])
    df_risultante = classi.sum()
    df_risultante["TOT.PEZZI.CLASSE"] = df_risultante.sum(axis=1)
    totale = (df_risultante[' [Bianca (0,45)]'] * 0.45) + (df_risultante[' [Rossa (1,00)]'] * 1) + (df_risultante[' [Margherita (1,00)]'] * 1 + (df_risultante[' [Marinara (1,00)]'] * 1) + (df_risultante[' [Patate (1,00)]'] * 1) + (df_risultante[' [Funghi Rossa (1,00)]'] * 1) + (df_risultante[' [Crostino (1,00)]'] * 1) + (df_risultante[' [Ripena Mortadella (1,00)]'] * 1) + (df_risultante[' [Ripiena Salame (1,00)]'] * 1) + (df_risultante[' [Ripiena Cotto (1,00)]'] * 1) + (df_risultante[' [Ripena Prosciutto (1,00)]'] * 1))
    df_risultante["TOT.PREZZO.CLASSE"] = totale
    # Aggiunta colonne pizza totale in base al tipo
    lista_colonne = df_risultante.columns
    for column in lista_colonne[:-1]:
        df_risultante[f"TOT.{column}"] = df_risultante[column].sum()
    df_risultante.rename(columns={"TOT.TOT.PEZZI.CLASSE":"TOT.PEZZI.ORDINE"},  inplace=True)
    righe_colonne = df_risultante.shape #tupla(n_righe,ncolonne)
    return df_risultante, righe_colonne

#directory ordini
def new_directory_ordini(percorso, data_ora):
    new_directory = f"{percorso}/ordini_{data_ora}"
    print(new_directory)
    if not os.path.exists(new_directory):
        nome_d = "ordini_" + str(data_ora)
        os.mkdir(f"{percorso}/{nome_d}")
    else:
        print("cartella esistente")
    return new_directory

#Directory Archivio
def archivia(lista_excel, percorso):
        data = data_ora[:10]
        day  = data_ora[8:10]
        h = data_ora[11:13]
        archivio_dir = percorso + "/archivio" + data
        if not os.path.exists(archivio_dir):
            os.mkdir(archivio_dir)
            for file in lista_excel:
                file = f"{percorso}/{file}"
                with open(file, "rb") as f:
                    old_file = f.read()
                with open(archivio_dir+f"/{day}-{h}file.xlsx", "wb") as f_arch:
                    f_arch.write(old_file)
                    os.remove(f"{percorso}/{lista_excel[0]}")
        else:
            for file in lista_excel:
                file = f"{percorso}/{file}"
                with open(file, "rb") as f:
                    old_file = f.read()
                with open(archivio_dir+f"/{day}-{h}file.xlsx", "wb") as f_arch:
                    f_arch.write(old_file)
                    f.close()
                    os.remove(f"{percorso}/{lista_excel[0]}")

#File elaborato di output
def gestisci_file(new_directory, df_risultante, data_ora):
    h_m = data_ora[11:16]
    nome_ordine = f"Ordine_delle_{h_m}"
    if not os.path.exists(new_directory + "/" + nome_ordine):
        df_risultante.to_excel(f"{new_directory}/{nome_ordine}.xlsx",float_format="%.2f")
        file_output = f"{new_directory}/{nome_ordine}.xlsx"
        return file_output
    else:
        print("File esistente")

#Formatta file output
def formatta_excel_output(file_output, righe_colonne):
    wb = load_workbook(filename=file_output)
    sheet = wb["Sheet1"]
    lista_colonne = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
                     "U", "V", "W", "X", "Y", "Z", "AA"]
    lista_colonne_rosse = ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA"]
    lista_colonne_nere = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    lista_colonne_valori = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T","U", "V", "W", "X", "Y", "Z", "AA"]
    pattern_fill = PatternFill(patternType='solid', fgColor="00C0C0C0")
    riga_intestazione = 1
    riga_dinamica = 2
    ft_black = Font(color="00000000", bold=True)
    ft_red = Font(color="00FF0000",bold=True)
    for colonna in lista_colonne:
        sheet.column_dimensions[colonna].width = 28
        if colonna in lista_colonne_valori:
            for riga in range (1, righe_colonne[0]+1,2):
                sheet[colonna + str(riga+1)].fill = pattern_fill
                riga_dinamica += 2
                if riga_dinamica == int(righe_colonne[0]-1):
                    riga_dinamica = 2
        if colonna in lista_colonne_rosse:
            sheet[colonna + str(riga_intestazione)].font = ft_red
        if colonna in lista_colonne_nere:
            sheet[colonna + str(riga_intestazione)].font = ft_black
    wb.save(filename=file_output)

#Invio Email
def invia_mail(new_directory, file_output):
    messaggio = EmailMessage()
    messaggio["Subject"] = "Ordine pizza"
    messaggio["From"] = "inviopizza@gmail.com"
    messaggio["To"] = "simone-tempesta@alice.it"
    with open(f"{file_output}", "rb") as f:
        file = f.read()
        messaggio.add_attachment(file, maintype='application', subtype='xlsx', filename='ordine.xlsx')
    print(messaggio.get_content_type())
    email = smtplib.SMTP("smtp.gmail.com", 587)
    email.ehlo()
    email.starttls()
    email.ehlo()
    email.login("inviopizza@gmail.com", "ForceisBig1977")
    email.send_message(messaggio)
    email.quit()

#Variabili
data_ora = datetime.now().strftime('%Y-%m-%d %H#%M#%S')
#percorso = str(os.getcwd())

def main(percorso):
    #MAIN
    #Elaborazione Exel
    l_directory = leggi_directory(percorso)
    lista_exel = estrai_file(l_directory)
    df_elaborato = elaborazione_exel(lista_exel, percorso)
    df_risultante, righe_colonne = elabora_df_elab(df_elaborato)

    #File e Directory
    #new_directory = percorso + "\\ordini_" + data_ora
    new_directory = new_directory_ordini(percorso, data_ora)
    file_output = gestisci_file(new_directory, df_risultante, data_ora)
    if file_output != None:
        formatta_excel_output(file_output, righe_colonne)
    return new_directory, lista_exel, file_output

if __name__ == "__main__":
    main()


